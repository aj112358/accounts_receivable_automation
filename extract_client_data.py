# This program extracts data from the client data.
# Created By: AJ Singh
# Date: Jan 10, 2021

from openpyxl import load_workbook
from calendar import month_name
from tabulate import tabulate

MONTHS = list(month_name)


# Opens Excel file containing client data.
# @params file_path: Path to main Excel file.
# @return: Excel worksheet.
def open_file(file_path: str):
    """Open Excel worksheet containing client data."""

    file = load_workbook(file_path, data_only=True)
    worksheet = file.active

    return worksheet


# Extract column headers from Excel sheet.
# @param worksheet: Excel worksheet.
# @return: List of active column headers.
# NOTE: Unused function as of now (Jan 12, 2021).
def get_column_headers(worksheet) -> list:
    """Get list of column headers from Excel sheet."""

    num_columns = worksheet.max_column
    headers = []
    for j in range(1, num_columns+1):
        cell_obj = worksheet.cell(row=1, column=j)
        headers.append(cell_obj.value)

    return headers


# Extract client names from Excel sheet.
# @param worksheet: Excel worksheet.
# @return: List of client names.
def get_client_names(worksheet) -> list:
    """Get list of client names from Excel worksheet."""

    num_rows = worksheet.max_row
    names = []
    for i in range(2, num_rows+1):
        cell_obj = worksheet.cell(row=i, column=1)
        if cell_obj.value not in names:
            names.append(cell_obj.value)

    return names


# Determine which clients have amounts owing.
# @param month: Specified month to check.
# @param names: List of client names.
# @param worksheet: Active Excel worksheet.
# @return: Dictionary of clients with amount owing.
def get_amounts_owing(month: str, names: list, worksheet) -> dict:
    """Compute amounts owing for each client."""

    # Dictionary values will have structure: [ (date, amount) ]
    amounts_owing = {name: [] for name in names}
    num_rows = worksheet.max_row

    for i in range(2, num_rows+1):
        paid = worksheet.cell(row=i, column=10).value
        month_num = worksheet.cell(row=i, column=4).value.month

        if not paid and month_num == MONTHS.index(month):
            name = worksheet.cell(row=i, column=1).value
            date = worksheet.cell(row=i, column=4).value
            owing = worksheet.cell(row=i, column=9).value

            amounts_owing[name].append((date, owing))

    return amounts_owing


# Extract client phone numbers from Excel worksheet.
# @params worksheet: Excel worksheet containing client data.
# @params names: List containing names of all clients.
# @return: Dictionary containing phone number of each client
def get_phone_nums(worksheet, names: list) -> dict:
    """Extract phone numbers of each client into a dictionary."""

    num_rows = worksheet.max_row
    phone = {name: "" for name in names}

    for i in range(2, num_rows+1):
        name = worksheet.cell(row=i, column=1).value
        number = worksheet.cell(row=i, column=12).value

        phone[name] = number

    return phone


# Print results of amounts owing to terminal.
# @param results: Dictionary containing clients' amounts owing.
# @param month: Specified month to check.
# @return: None.
def print_results(results: dict, month: str) -> None:
    """Print amounts owing for each client to the terminal."""

    table_headers = ["Name", "Owing"]
    data = []

    print(f"\nHere are the amounts owing for {month}:")
    for name in results:
        if results[name]:  # If there is an amount owing.
            totals = list(zip(*results[name]))[1]
            total = sum(totals)
            data.append([name, "${0:.2f}".format(total)])
        else:
            data.append([name, "All good!"])

    print("\n", tabulate(data, table_headers))


# Creates text file containing name/amount owing/phone number of client.
# @param results: Dictionary containing clients' amounts owing.
# @param month: Specified month to check.
# @param phone_nums: Dictionary containing clients' phone numbers.
# @return: None.
def save_results(results: dict, month: str, phone_nums: dict) -> None:
    """Create text file containing a client's name, amount owing, and phone number."""

    with open(f"{month}.txt", mode='w') as out_file:
        for name in results:
            if results[name]:  # Only record non-zero amounts owing.
                totals = list(zip(*results[name]))[1]
                total = sum(totals)
                phone = phone_nums[name]
                out_file.write("{0},${1:.2f},{2}\n".format(name, total, phone))


# Main function to extract client data from Excel sheet and
# store amounts owing in a text file.
# @param path: Path to main Excel file.
# @param month: Specified month to check.
# @return: None.
def extract_client_data(file_path: str, month: str) -> None:
    """Get amounts owing for each client, print results to terminal, and save to text file."""

    worksheet = open_file(file_path)

    names = get_client_names(worksheet)
    phone_nums = get_phone_nums(worksheet, names)

    results = get_amounts_owing(month, names, worksheet)

    print_results(results, month)
    save_results(results, month, phone_nums)


if __name__ == "__main__":
    path = r"C:\Users\AJ\Desktop\accounts_receivable_automation\client_data_sample.xlsx"
    month_name = input("\nEnter full name of month: ").capitalize()
    while month_name not in MONTHS:
        print("Invalid month!")
        month_name = input("\nEnter full name of month: ").capitalize()
    extract_client_data(path, month_name)
